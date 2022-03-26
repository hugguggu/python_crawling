from pickle import NONE
from webbrowser import get
import openpyxl
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.action_chains import ActionChains

# # url = 'https://www.etf.com/modal_forms/nojs/login?destination=//'
# url = 'https://www.etf.com'
# browser = webdriver.Chrome('C:/chromedriver.exe')
# browser.get(url)
# # time.sleep(5)
# wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'loginLnk')))
# print(browser.find_element_by_id('edit-name').text)


options = webdriver.ChromeOptions()
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
options.add_argument("lang=ko_KR")
options.add_argument("user-agent=Mozilla/5.0 (M acintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")
options.add_argument("Accept=text/html,application/xhtml+xml,application/xml;\q=0.9,imgwebp,*/*;q=0.8")
# options.add_argument("proxy-server=localhost:8080")
driver = webdriver.Chrome('./chromedriver.exe', options=options)
# driver = webdriver.Chrome('C:/chromedriver.exe')

# url = 'https://www.etf.com/etfanalytics/etf-finder'
url = 'https://www.etf.com/modal_forms/nojs/login?destination=/'
# url = 'https://www.etf.com'
driver.get(url)
driver.maximize_window()

# element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="table-tabs"]/li[7]')))
# element.send_keys(Keys.ENTER)

res = 0
while res < 1:
        try:
                time.sleep(2)
                element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "edit-name")))
                element.click()
                element.send_keys('hugguggu@gmail.com')

                time.sleep(1)
                element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "edit-pass")))
                element.click()
                element.send_keys('Rodroddl01!')

                time.sleep(2)
                element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "edit-submit")))
                element.click()
                print('log in ok\n')
                res = 1
        except:
                res = 0
                print('log in fail\n')

res = 0

while res < 1:
        try:

                div_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, 'menu-nav')))
                hover = ActionChains(driver).move_to_element(div_element)
                hover.perform()

                # driver.execute_script("window.scrollTo(0, Y)") 
                # //*[@id="menu-nav"]/li[1]/span/div
                button = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="menu-nav"]/li[1]/span/div')))
                hover = ActionChains(driver).move_to_element(button)
                hover.perform()
                # //*[@id="menu-nav"]/li[1]/ul/li[1]/a
                element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="menu-nav"]/li[1]/ul/li[1]/a')))
                hover = ActionChains(driver).move_to_element(element).click()
                hover.perform()

                # time.sleep(3)

                div_element = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.ID, 'finder-root')))
                hover = ActionChains(driver).move_to_element(div_element)
                hover.perform()

                res = 1
        except:
                res = 0

res = 0
while res < 1:
        try:
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'results_display')))
                hover = ActionChains(driver).move_to_element(element)
                hover.perform()
                driver.execute_script("arguments[0].scrollIntoView(true);", element)

                # /html/body/div[7]/section/div/div[3]/section/div/div/div/div/div[2]/section[2]/div[2]/section[2]/div[1]/div/div[3]/button
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="results_display"]/div/div[4]/button')))
                hover = ActionChains(driver).move_to_element(element).click()
                hover.perform()

                # time.sleep(3)
                div_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'table-tabs')))
                hover = ActionChains(driver).move_to_element(div_element)
                hover.perform()
                driver.execute_script("arguments[0].scrollIntoView(true);", div_element)

                # time.sleep(3)
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="table-tabs"]/li[7]')))
                hover = ActionChains(driver).move_to_element(element).click()
                hover.perform()
                # driver.execute_script("arguments[0].scrollIntoView(true);", element)
                
                # time.sleep(3)
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'more-btn')))
                hover = ActionChains(driver).move_to_element(element).click()
                hover.perform()
                res = 1
        except:
                res = 0
                print('retry\n')



# if driver.find_element_by_xpath('//input[@type="checkbox"]').get_attribute('checked'):  
#         driver.find_elements_by_class_name('uiInputLabelLabel')[-1].click()   


# open custom tab
div_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'finderCustomTabOptions')))
hover = ActionChains(driver).move_to_element(div_element)
hover.perform()

div_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="finderCustomTabOptions"]/div[2]')))
hover = ActionChains(driver).move_to_element(div_element)
hover.perform()
# open custom tab end


# checking
value = driver.find_elements(By.CLASS_NAME, 'form-check')
print(len(value))
# /html/body/div[7]/section/div/div[3]/section/div/div/div/div/div[2]/section[2]/div[2]/div[2]/div/section/div[2]/div[1]/span/label/input

for idx in range (1, 59, 1):
        element = driver.find_element(By.XPATH, f'//*[@id="finderCustomTabOptions"]/div[2]/div[{idx}]/span/label')
        checked = driver.find_element(By.XPATH, f'//*[@id="finderCustomTabOptions"]/div[2]/div[{idx}]/span/label/input')
        if checked.get_attribute('checked') :
                print( ' checked')        
        else:
                # //*[@id="finderCustomTabOptions"]/div[2]/div[1]/span/label
                # div_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="finderCustomTabOptions"]/div[2]/div[4]/span/label')))
                hover = ActionChains(driver).move_to_element(element).click()
                hover.perform()
                print(' change')        

# checking end

# close custom tab
div_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="finderCustomTabOptions"]/div[3]')))
hover = ActionChains(driver).move_to_element(div_element)
hover.perform()

div_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="finderCustomTabOptions"]/div[3]/button[2]')))
hover = ActionChains(driver).move_to_element(div_element).click()
hover.perform()
# close custom tab end


# sorting
div_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="results"]/div/div[2]')))
hover = ActionChains(driver).move_to_element(div_element)
hover.perform()
driver.execute_script("arguments[0].scrollIntoView(true);", div_element)

div_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'custom_LaunchDate')))
hover = ActionChains(driver).move_to_element(div_element).click()
hover.perform()
time.sleep(1)
hover = ActionChains(driver).move_to_element(div_element).click()
hover.perform()
# sorting end

# write excel
element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'finderTable')))

thead = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.TAG_NAME, 'thead')))
header = thead.find_elements(By.TAG_NAME, 'label')

tbody = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.TAG_NAME, 'tbody')))
rows = tbody.find_elements(By.TAG_NAME, 'tr')

# for index, value in enumerate(rows):
#     body=value.find_elements_by_tag_name("td")[1]
#     print(body.text)


print(' writing')
# 엑셀 파일 생성
workbook = openpyxl.Workbook()

# 엑셀 워크시트 생성
# sheet = workbook.create_sheet("test")
sheet = workbook.worksheets[0]
ft = Font(bold=True)
for idx, tag in enumerate(header):
        sheet.cell(row = 1, column = idx + 1).value = tag.text
        sheet.cell(row = 1, column = idx + 1).font = ft

for row_idx, row_val in enumerate(rows):
        col = row_val.find_elements_by_tag_name("td")
        print(f'row : {row_idx + 1}')
        for td_idx, col_val in enumerate(col):
                # print(col_val.text, end=" ")
                # sheet.cell(row = row_idx + 2, column = td_idx + 1).value = col_val.text
                sheet.cell(row = row_idx + 2, column = td_idx + 1).value = col_val.text
                if td_idx == 0:
                        sheet.cell(row = row_idx + 2, column = td_idx + 1).hyperlink = 'http://www.etf.com/{}'.format(col_val.text)
               
#         # print("")
        
workbook.save(r'etf_com.xlsx')

driver.quit()

