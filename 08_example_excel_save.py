import openpyxl

# 엑셀 파일 생성
workbook = openpyxl.Workbook()

# 엑셀 워크시트 생성
sheet = workbook.create_sheet("test")

# 데이터 추가
sheet['A1'] = '참가번호'
sheet['B1'] = '성명'

sheet['A2'] = '1'
sheet['B2'] = 'ABC'

sheet.cell(row=1, column=40).value = 'TESTSSS'

workbook.save(r'test_excel.xlsx')



